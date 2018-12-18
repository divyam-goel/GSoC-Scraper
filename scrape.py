import sys
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

wb = Workbook()

if len(sys.argv)>1:
	category = '_'.join(sys.argv[1:])
	ws = wb.create_sheet(title=category)
	url = "https://summerofcode.withgoogle.com/archive/2018/organizations/?category="+category

else:
	ws = wb.create_sheet(title="Organizations")
	url = "https://summerofcode.withgoogle.com/archive/2018/organizations/"

font_name = Font(size=13)

alignment_name = Alignment(wrap_text=True,
                           vertical="center", horizontal="left")
alignment = Alignment(wrap_text=True, shrink_to_fit=True,
                      vertical="center", horizontal="left")

ws["A1"] = "Name"
ws.column_dimensions["A"].width = 30
ws["B1"] = "Tagline"
ws.column_dimensions["B"].width = 40
ws["C1"] = "Tech/Topic"
ws.column_dimensions["C"].width = 40
ws["D1"] = "Project Count"
ws.column_dimensions["D"].width = 10
ws["E1"] = "Link"
ws.column_dimensions["E"].width = 65

ws.row_dimensions[1].height = 30
ws.row_dimensions[1].font = Font(size=16, bold=True)
ws.row_dimensions[1].alignment = Alignment(
    vertical="center", horizontal="center", shrink_to_fit=True)

response = requests.get(url)
html = response.content

soup = BeautifulSoup(html, features="html.parser")
orgs = soup.findAll('li', attrs={'class': 'organization-card__container'})

default = "https://summerofcode.withgoogle.com"
for index, org in enumerate(orgs):
    
    name = org['aria-label']
    link = org.find('a', attrs={'class': 'organization-card__link'})
    link = default + link['href']
    tagline = org.find(
        'div', attrs={'class': 'organization-card__tagline'}).string

    response = requests.get(link)
    html = response.content
    soup = BeautifulSoup(html, features="html.parser")

    technologies = soup.findAll(
        'li', attrs={"class": "organization__tag organization__tag--technology"})
    tech_list = ''
    for tech in technologies:
        tech_list = tech_list + tech.string + ", "

    topics = soup.findAll(
        'li', attrs={"class": "organization__tag organization__tag--topic"})
    topic_list = ''
    for topic in topics:
        topic_list = topic_list + topic.string + ", "

    tech_topic = "Tech: " + tech_list + '\n' + "Topic: " + topic_list

    projects = soup.find("ul", attrs={"class": "project-list-container"})
    project_count = len(projects.findAll("li"))

    ws["A%d" % (index + 2)] = name
    ws["A%d" % (index + 2)].alignment = alignment_name
    ws["A%d" % (index + 2)].font = font_name

    ws["B%d" % (index + 2)] = tagline
    ws["B%d" % (index + 2)].alignment = alignment

    ws["C%d" % (index + 2)] = tech_topic
    ws["C%d" % (index + 2)].alignment = alignment

    ws["D%d" % (index + 2)] = project_count
    ws["D%d" % (index + 2)].alignment = alignment

    ws["E%d" % (index + 2)] = link
    ws["E%d" % (index + 2)].alignment = alignment

    ws.row_dimensions[index + 2].height = 55

wb.save(filename="GSoC-Organizations.xlsx")

# Test
